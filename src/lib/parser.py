from openpyxl import load_workbook
import json


def parse(path):
    wb = load_workbook(path)
    sheets = wb.sheetnames
    link_dict = {}
    for sheet in sheets:
        link_dict[sheet] = []

        for row in wb[sheet].rows:
            link_dict[sheet].append((row[0].value))
        del(link_dict[sheet][0])

    with open('links.json', "w") as file:
        file.write(json.dumps((link_dict), sort_keys=True, indent=2))

    